import telebot
from telebot import types
import os
from openpyxl import Workbook, load_workbook

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

bot = telebot.TeleBot(BOT_TOKEN)

# 🚘 Mercedes модельдері (баға + сипаттама)
cars = {
    "A-Class": {"price": 20000000, "desc": "Компакт класс, жастарға арналған стильді седан"},
    "C-Class": {"price": 25000000, "desc": "Динамика мен статус үйлесімі. 2.0L Turbo, 204 ат күші, 0-100: 7.3с"},
    "CLA": {"price": 28000000, "desc": "Купе стиліндегі премиум автомобиль, динамикасы жоғары"},
    "E-Class": {"price": 35000000, "desc": "Бизнес класс стандарты. Комфорт пен технология балансы"},
    "GLA": {"price": 32000000, "desc": "Компакт SUV, қала және жолға жарамды"},
    "GLB": {"price": 45000000, "desc": "7 орындық шағын люкс SUV"},
    "S-Class": {"price": 60000000, "desc": "Люкс сегмент көшбасшысы. Максималды комфорт"},
    "G-Class": {"price": 120000000, "desc": "Күш пен статус символы. Толық жетекті премиум SUV"},
    "EQE": {"price": 80000000, "desc": "Электрлік седан, премиум технологиялар"},
    "EQS": {"price": 140000000, "desc": "Электрлік люкс седан, максималды комфорт және инновация"}
}

cities = ["Алматы", "Астана", "Шымкент"]

# Excel жасау
if not os.path.exists("clients.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Phone", "City", "Car", "Date"])
    wb.save("clients.xlsx")

# 🟢 START
@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("🚘 Модельдер", "💰 Бюджет бойынша таңдау")
    markup.add("📋 Тест-драйв", "🧩 Сізге қай тип көлік керек?")

    bot.send_message(
        message.chat.id,
        "✨ Mercedes-Benz ресми цифрлық менеджеріне қош келдіңіз.\nПремиум әлемге қадам жасаңыз.",
        reply_markup=markup
    )

# Модельдер көрсету
@bot.message_handler(func=lambda m: m.text == "🚘 Модельдер")
def show_cars(message):
    text = "🚘 Қол жетімді модельдер:\n\n"
    for name, info in cars.items():
        text += f"{name} — {info['price']:,} ₸\n{info['desc']}\n\n"
    bot.send_message(message.chat.id, text)

# Бюджет бойынша кеңес
@bot.message_handler(func=lambda m: m.text == "💰 Бюджет бойынша таңдау")
def ask_budget(message):
    bot.send_message(message.chat.id, "Бюджетіңізді жазыңыз (мысалы: 40000000):")

@bot.message_handler(func=lambda m: m.text.isdigit())
def recommend_car(message):
    budget = int(message.text)
    recommended = None

    for name, info in cars.items():
        if budget >= info["price"]:
            recommended = name

    if recommended:
        bot.send_message(
            message.chat.id,
            f"💎 Сізге {recommended} ұсынылады.\n{cars[recommended]['desc']}\n\nТест-драйв ұйымдастырайық па?"
        )
    else:
        bot.send_message(message.chat.id, "Өкінішке орай бұл бюджетке модель жоқ.")

# Клиент типі бойынша ұсыныс
@bot.message_handler(func=lambda m: m.text == "🧩 Сізге қай тип көлік керек?")
def recommend_type(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Отбасы", "Жастар", "Бизнес", "Электрлік", "Жолшы")
    bot.send_message(message.chat.id, "Сізге қандай көлік керек?", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text in ["Отбасы", "Жастар", "Бизнес", "Электрлік", "Жолшы"])
def suggest_models(message):
    type_choice = message.text
    suggestions = []

    if type_choice == "Отбасы":
        suggestions = ["GLB", "GLA", "C-Class"]
    elif type_choice == "Жастар":
        suggestions = ["A-Class", "CLA"]
    elif type_choice == "Бизнес":
        suggestions = ["E-Class", "S-Class"]
    elif type_choice == "Электрлік":
        suggestions = ["EQE", "EQS"]
    elif type_choice == "Жолшы":
        suggestions = ["G-Class"]

    text = f"💎 {type_choice} клиенттерге келесі модельдер ұсынылады:\n\n"
    for model in suggestions:
        info = cars[model]
        text += f"{model} — {info['price']:,} ₸\n{info['desc']}\n\n"

    bot.send_message(message.chat.id, text)

# Тест-драйв
@bot.message_handler(func=lambda m: m.text == "📋 Тест-драйв")
def test_drive(message):
    bot.send_message(message.chat.id, "Атыңыз:")
    bot.register_next_step_handler(message, get_name)

def get_name(message):
    name = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for city in cities:
        markup.add(city)
    bot.send_message(message.chat.id, "Қай қалада тест-драйв?", reply_markup=markup)
    bot.register_next_step_handler(message, get_city, name)

def get_city(message, name):
    city = message.text
    bot.send_message(message.chat.id, "Телефон:")
    bot.register_next_step_handler(message, get_phone, name, city)

def get_phone(message, name, city):
    phone = message.text
    bot.send_message(message.chat.id, "Қай модель?")
    bot.register_next_step_handler(message, get_car, name, city, phone)

def get_car(message, name, city, phone):
    car = message.text
    bot.send_message(message.chat.id, "Күні (15.02.2026 15:00):")
    bot.register_next_step_handler(message, save_data, name, city, phone, car)

def save_data(message, name, city, phone, car):
    date = message.text
    wb = load_workbook("clients.xlsx")
    ws = wb.active
    ws.append([name, phone, city, car, date])
    wb.save("clients.xlsx")
    bot.send_message(message.chat.id, "✅ Сұраныс қабылданды. Біз сізбен байланысамыз.")

# Статистика
@bot.message_handler(commands=['stats'])
def stats(message):
    if message.chat.id != ADMIN_ID:
        return

    wb = load_workbook("clients.xlsx")
    ws = wb.active
    counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        car = row[3]
        counts[car] = counts.get(car, 0) + 1

    text = "📊 Статистика:\n\n"
    for car, count in counts.items():
        text += f"{car} — {count} заявка\n"

    bot.send_message(message.chat.id, text)

# Excel жүктеу
@bot.message_handler(commands=['clients'])
def send_excel(message):
    if message.chat.id == ADMIN_ID:
        with open("clients.xlsx", "rb") as f:
            bot.send_document(message.chat.id, f)

print("Bot running...")
bot.infinity_polling()
